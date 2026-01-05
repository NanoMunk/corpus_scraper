import random
import re
import openpyxl
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

BASE_DIR = Path(__file__).resolve().parent

corpus = "https://corpora.uni-leipzig.de/en/res?corpusId=fin-fi_web_2019&word="
# yle_corpus_2022 = "https://corpora.uni-leipzig.de/en/res?corpusId=fin_news_2022&word="
# implemented later

# Search file location
data_path = BASE_DIR/"data"/"search_file.xlsx"

# workbook object is created
wb_obj = openpyxl.load_workbook(data_path)

# Get workbook
sheet_obj = wb_obj.active


search_word_list = []

max_row_num = sheet_obj.max_row

for i in range(1, max_row_num + 1):
	cell_obj = sheet_obj.cell(row=i, column=3)
	search_word_list.append(cell_obj.value)

print(f'Number of words: {len(search_word_list)}')
print("\n")

driver = webdriver.Firefox()

for i in range(0,len(search_word_list)):
	search_word = str(search_word_list[i])

	driver.get(str(corpus + search_word))

	driver.implicitly_wait(30)

	try:

		if driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/p").text == "Word not found":

			print("\nNo Results found")
			print("----------------------------------------\n")

			continue

	except:

		pass

	try:

		example_box = driver.find_element(By.XPATH, "/html/body/div[1]/div[1]/div[1]/div[1]/div[4]/div[1]/div[2]/ul[1]").text

		pattern = re.compile(r"^(.*?)\s?\(", re.MULTILINE)
	
		examples_no_source = pattern.findall(example_box)

	except:

		continue

	example_cell = sheet_obj["A" + str((search_word_list.index(search_word)+1))]

	example_cell.value = random.choice(examples_no_source)
	print(f'Example for {search_word}: | {example_cell.value} |')
	print("----------------------------------------\n")

	wb_obj.save(data_path)